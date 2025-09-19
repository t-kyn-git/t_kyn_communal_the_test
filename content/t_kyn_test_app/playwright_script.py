import os
import sys
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# --- 設定項目 ---
BASE_URL = "http://127.0.0.1:18000"
LOGIN_URL = f"{BASE_URL}/admin/login/" # ログインページのURLに合わせて変更してください
USERNAME = "YOUR_USERNAME"  # Djangoのログインユーザー名に変更
PASSWORD = "YOUR_PASSWORD"  # Djangoのログインパスワードに変更

# スクリーンショットを撮るページのパス
URL_PATHS_TO_CAPTURE = {
    "album": "/album/",
    "chat": "/chat/",
    "todolist": "/todolist/",
    "location": "/location/",
}

# --- メイン処理 ---
def main():
    # スクリーンショットを保存する一時ディレクトリを作成
    screenshots_dir = "screenshots"
    os.makedirs(screenshots_dir, exist_ok=True)

    with sync_playwright() as p:
        try:
            browser = p.chromium.launch()
            page = browser.new_page()

            # 1. ログイン処理
            print(f"Navigating to login page: {LOGIN_URL}")
            page.goto(LOGIN_URL)
            
            # ログインフォームに入力して送信（セレクタは実際のHTMLに合わせて調整してください）
            page.fill("input[name='username']", USERNAME)
            page.fill("input[name='password']", PASSWORD)
            page.click("input[type='submit']")
            
            # ログイン後のページ読み込みを待機
            page.wait_for_load_state("networkidle")
            print("Login successful.")

            # 2. 各ページでスクリーンショットを撮影
            screenshot_files = {}
            for name, path in URL_PATHS_TO_CAPTURE.items():
                target_url = f"{BASE_URL}{path}"
                print(f"Capturing screenshot of: {target_url}")
                page.goto(target_url, wait_until="networkidle")
                
                # 少し待機して描画を安定させる
                page.wait_for_timeout(1000)

                filepath = os.path.join(screenshots_dir, f"{name}.png")
                page.screenshot(path=filepath, full_page=True)
                screenshot_files[name] = filepath
                print(f"Screenshot saved to: {filepath}")

            browser.close()

            # 3. Excelファイルに画像を貼り付け
            print("Creating Excel file...")
            wb = Workbook()
            ws = wb.active
            ws.title = "ScreenCaptures"
            ws['A1'] = "Page"
            ws['B1'] = "Screenshot"

            for i, (name, img_path) in enumerate(screenshot_files.items(), start=2):
                ws[f'A{i}'] = name
                img = Image(img_path)
                # 画像サイズを調整（任意）
                img.width = 640
                img.height = 480
                ws.add_image(img, f'B{i}')

            output_excel_file = "django_captures.xlsx"
            wb.save(output_excel_file)
            print(f"Excel file created successfully: {output_excel_file}")

        except Exception as e:
            print(f"An error occurred: {e}", file=sys.stderr)
            # エラー発生時にもスクリーンショットを撮っておくとデバッグに役立つ
            page.screenshot(path="error_screenshot.png")
            sys.exit(1)


if __name__ == "__main__":
    main()